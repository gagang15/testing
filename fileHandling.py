import openpyxl

path="D:\demo.xlsx"
workbook=openpyxl.load_workbook(path)
#sheet1=workbook.active
#sheet1=workbook.get_sheet_by_name("Sheet2")
sheet=workbook["Sheet1"]
rows=sheet.max_row
cols=sheet.max_column
print(rows,cols)
# sheet1.cell(row=1,column=1).value="Hello"
n=6
for i in range(1,n):
    for j in range(i):
         sheet.cell(row=i,column=j+1).value="*"
         data=sheet.cell(row=i,column=j+1).value
         print(data,end="  ")
    print(" ")

workbook.save(path)