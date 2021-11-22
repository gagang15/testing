import openpyxl

class DataReader:
    @staticmethod
    def readDataforlogin():
        path = "D:\demo.xlsx"
        workbook = openpyxl.load_workbook(path)
        sheet = workbook["Sheet1"]
        rows = sheet.max_row
        cols = sheet.max_column
        for i in range(1, rows + 1):
            data= sheet.cell(row=i,column=1).value
            if data == "tc_004":
                userName = sheet.cell(row=i, column=2).value
                passWord = sheet.cell(row=i, column=3).value
                return (userName, passWord)



# ob=DataReader()
# username,password=ob.readDataforlogin()
# print(username,password)



if __name__=="__main__":
    DataReader.readDataforlogin()