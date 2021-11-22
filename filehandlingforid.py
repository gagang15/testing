import openpyxl
from selenium import webdriver
path="D:\demo.xlsx"
workbook=openpyxl.load_workbook(path)
#sheet1=workbook.active
#sheet1=workbook.get_sheet_by_name("Sheet2")
sheet=workbook["Sheet1"]
rows=sheet.max_row
cols=sheet.max_column
print(rows)
# sheet1.cell(row=1,column=1).value="Hello"
n=6
for i in range(1,rows+1):
    for j in range(cols):
         data=sheet.cell(row=i,column=j+1).value
         if data=="tc_004":
             k=j
             print(sheet.cell(row=i,column=j+1).value,end=" ")
             userName=sheet.cell(row=i,column=k+2).value
             passWord=sheet.cell(row=i, column=k+3).value


driver = webdriver.Chrome(executable_path="E:\\Automation project demo\\browser drive\chromedriver.exe")
driver.maximize_window()
driver.get("http://demo.guru99.com/v2/")
if driver.find_element_by_name("uid").is_displayed():
    if driver.find_element_by_name("uid").is_enabled():
        driver.find_element_by_name("uid").send_keys(userName)  # entering user id
    else:
        print("user name is disabled")

else:

    print("user name is not displyed ")

if driver.find_element_by_name("password").is_displayed():
    if driver.find_element_by_name("password").is_enabled():
        driver.find_element_by_name("password").send_keys(passWord)  # entering password id
    else:
        print("Password  is disabled")

else:

    print("Password is not displyed ")

driver.find_element_by_name("btnLogin").click()




workbook.save(path)