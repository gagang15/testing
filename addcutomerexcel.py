import openpyxl
import time
from selenium import webdriver
path="D:\demo.xlsx"
workbook=openpyxl.load_workbook(path)

sheet=workbook["Sheet3"]
rows=sheet.max_row
cols=sheet.max_column

dict1={}
for i in range(1,rows+1):
    data = sheet.cell(row=i, column=1).value
    if data == "tc_004":
        for j in range(2,cols+1):
            dict1[sheet.cell(row=1,column=j).value]=str(sheet.cell(row=i,column=j).value)

print(dict1)

driver=webdriver.Chrome(executable_path="E:\\Automation project demo\\browser drive\chromedriver.exe")
driver.get("http://demo.guru99.com/v2/")
driver.find_element_by_name("uid").send_keys("mngr346462") #entering user id
driver.find_element_by_name("password").send_keys("hahamyz") # entering Password
driver.find_element_by_name("btnLogin").click()
driver.find_element_by_xpath("/html/body/div[3]/div/ul/li[2]/a").click()
driver.find_element_by_name("name").send_keys(dict1['name'])
driver.find_element_by_name("rad1").click()
driver.find_element_by_name("dob").send_keys("09-09-1991")
driver.find_element_by_name("addr").send_keys(dict1['addr'])
driver.find_element_by_name("city").send_keys(dict1['city'])
driver.find_element_by_name("state").send_keys(dict1['State'])
driver.find_element_by_name("pinno").send_keys(dict1['pin'])
driver.find_element_by_name("telephoneno").send_keys(dict1['mobNo'])
driver.find_element_by_name("emailid").send_keys(dict1['emailId'])
driver.find_element_by_name("sub").click()
time.sleep(3)



