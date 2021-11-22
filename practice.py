import time


from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.remote.webelement import WebElement

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
# driver=webdriver.Chrome(executable_path="E:\\Automation project demo\\browser drive\chromedriver.exe")
#
# driver.maximize_window()
#
# driver.get("http://demo.guru99.com/telecom/assigntariffplantocustomer.php")
# # # driver.execute_script('document.getElementsByTagName("video")[0].play()')
# driver.find_element_by_name('customer_id').send_keys('283545')
# driver.find_element_by_name('submit').click()
# table1=driver.find_element_by_xpath('//*[@id="main"]/div/div/table')
# row=table1.find_elements_by_tag_name('tr')
# # for item in row:
# #     x = int(i)
# #     plan=driver.find_element_by_xpath('//*[@id="main"]/div/div/table/tbody['+str(x)+']/tr/td[2]')
# #     p
# #     i=int(i)+1
# for i in range(len(row)):
#     i=i+1
#     if i==len(row):
#         break
#     plan = driver.find_element_by_xpath('//*[@id="main"]/div/div/table/tbody[' + str(i) + ']/tr/td[2]')
#     #print(plan.text)
#     if plan.text=="2222":
#         print('***')
#         result=driver.find_element_by_xpath('//*[@id="main"]/div/div/table/tbody[' + str(i) + ']/tr/td[1]')
#         print(result.text)
#

#driver.close()
#
#
# ############################################################
# #Conditional for new Customer
# def fnNewCustomer(**newcustomer):
#     driver.get("http://demo.guru99.com/v2/")
#     if driver.find_element_by_name("uid").is_displayed():
#         if driver.find_element_by_name("uid").is_enabled():
#             driver.find_element_by_name("uid").send_keys(newcustomer["userName"])  # entering user id
#         else:
#             print("user name is disabled")
#             return
#     else:
#
#         print("user name is not displyed ")
#         return
#     if driver.find_element_by_name("password").is_displayed():
#         if driver.find_element_by_name("password").is_enabled():
#             driver.find_element_by_name("password").send_keys(newcustomer['password'])  # entering user id
#         else:
#             print("Password  is disabled")
#             return
#     else:
#
#         print("Password is not displyed ")
#         return
#     driver.find_element_by_name("btnLogin").click()
#     driver.find_element_by_xpath("/html/body/div[3]/div/ul/li[2]/a").click()
#     if driver.find_element_by_name("name").is_displayed():
#         if driver.find_element_by_name("name").is_enabled():
#             driver.find_element_by_name("name").send_keys(newcustomer["name"])  # entering user id
#         else:
#             print(" name is disabled")
#             return
#     else:
#
#         print(" name is not displyed ")
#         return
#
#     driver.find_element_by_name("rad1").click()
#     if driver.find_element_by_name("dob").is_displayed():
#         if driver.find_element_by_name("dob").is_enabled():
#             driver.find_element_by_name("dob").send_keys(newcustomer["dob"])  # entering user id
#         else:
#             print(" dob is disabled")
#             return
#     else:
#
#         print(" dob is not displyed ")
#         return
#     if driver.find_element_by_name("addr").is_displayed():
#         if driver.find_element_by_name("addr").is_enabled():
#             driver.find_element_by_name("addr").send_keys(newcustomer["addres"])  # entering user id
#         else:
#             print(" addres is disabled")
#             return
#     else:
#
#         print(" addres is not displyed ")
#         return
#     if driver.find_element_by_name("city").is_displayed():
#         if driver.find_element_by_name("city").is_enabled():
#             driver.find_element_by_name("city").send_keys(newcustomer["city"])  # entering user id
#         else:
#             print(" city is disabled")
#             return
#     else:
#
#         print(" city is not displyed ")
#         return
#     if driver.find_element_by_name("state").is_displayed():
#         if driver.find_element_by_name("state").is_enabled():
#             driver.find_element_by_name("state").send_keys(newcustomer["state"])  # entering user id
#         else:
#             print(" state is disabled")
#             return
#     else:
#
#         print(" state is not displyed ")
#         return
#     if driver.find_element_by_name("pinno").is_displayed():
#         if driver.find_element_by_name("pinno").is_enabled():
#             driver.find_element_by_name("pinno").send_keys(newcustomer["pin"])  # entering user id
#         else:
#             print(" pin is disabled")
#             return
#     else:
#
#         print(" pin is not displyed ")
#         return
#     if driver.find_element_by_name("telephoneno").is_displayed():
#         if driver.find_element_by_name("telephoneno").is_enabled():
#             driver.find_element_by_name("telephoneno").send_keys(newcustomer["mobileNo"])  # entering user id
#         else:
#             print(" pin is disabled")
#             return
#     else:
#
#         print(" pin is not displyed ")
#         return
#     if driver.find_element_by_name("emailid").is_displayed():
#         if driver.find_element_by_name("emailid").is_enabled():
#             driver.find_element_by_name("emailid").send_keys(newcustomer["email"])  # entering user id
#         else:
#             print(" email is disabled")
#             return
#     else:
#
#         print(" email is not displyed ")
#         return
#
#     driver.find_element_by_name("sub").click()
#
#

#
# newcustomer={
#     "userName":"mngr346462",
#     "password":"hahamyz",
#     "name":"Pratik Tiwari",
#     "dob":"12-11-1995",
#     "addres":"Nandanvan near padav collage",
#     "city":"nagpur",
#     "state":"Maharashtra",
#     "pin":"440009",
#     "mobileNo":"7502311121",
#     "email":"prtik@gmail.com"
# }
# fnNewCustomer(**newcustomer)
#
#
import openpyxl
import time
from selenium import webdriver
path="D:\demo.xlsx"
workbook=openpyxl.load_workbook(path)
#sheet1=workbook.active
#sheet1=workbook.get_sheet_by_name("Sheet2")
sheet=workbook["Sheet3"]
rows=sheet.max_row
cols=sheet.max_column
#userName=sheet.cell(row=i,column=k+2).value
n=6

keys1=[]
dict1={}
for j in range(1,cols+1):
        keys1.append(sheet.cell(row=1,column=j).value)

for i in (keys1):
    dict1[i]=None
j=1
for i in dict1:
    dict1[sheet.cell(row=1,column=j).value]=sheet.cell(row=2,column=j).value
    j+=1
print(dict1)